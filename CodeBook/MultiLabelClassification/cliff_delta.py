import os
import csv
import tensorflow as tf
import sys
import time
import random
import pickle
import argparse
import pandas as pd
from openpyxl import Workbook, load_workbook
from joblib import dump, load
from sklearn import preprocessing
from sklearn.multioutput import MultiOutputClassifier
from cliffs_delta import cliffs_delta
from sklearn.decomposition import PCA
from sklearn.feature_selection import mutual_info_classif
from sklearn.metrics import f1_score
from sklearn.metrics import roc_auc_score
from sklearn.metrics import matthews_corrcoef
from sklearn import tree
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from sklearn.model_selection import train_test_split
from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF
from sklearn.tree import DecisionTreeClassifier, plot_tree, export_graphviz
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.impute import SimpleImputer


def label_encoding(raw_labels, labelEncoder=None):
    if labelEncoder is None:
        labelEncoder = preprocessing.LabelEncoder()
        labelEncoder = labelEncoder.fit(raw_labels)
    enc_label = labelEncoder.transform(raw_labels)
    # print('classes: {}'.format(', '.join(labelEncoder.classes_)))
    return labelEncoder, enc_label


def label_decoder(le, enc_label):
    return le.inverse_transform(enc_label)


def multi_label_binarizer(raw_labels, labelEncoder=None):
    if labelEncoder is None:
        labelEncoder = preprocessing.MultiLabelBinarizer()
        labelEncoder = labelEncoder.fit(raw_labels)
    enc_label = labelEncoder.transform(raw_labels)
    # print('classes: {}'.format(', '.join(labelEncoder.classes_)))
    return labelEncoder, enc_label


def train(x_train, y_train, n_features):
    clf = MultiOutputClassifier(KNeighborsClassifier(n_neighbors=n_features)).fit(x_train, y_train)
    return clf


def predict(clf, x_test):
    clf.predict(x_test)


def balance(df_all, FOCUS, n_sample=100):
    sampled = df_all[df_all[FOCUS] == "origin"].sample(n=n_sample, random_state=1)
    df_new = pd.concat([sampled, df_all[df_all[FOCUS] != "origin"]])
    print("Sample before: {}, after: {}".format(df_all.shape, df_new.shape))
    return df_new


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Multi-Lable Classifier Training.")
    parser.add_argument('--program_dir', '-pd', default='./RQ2', help='program path.')
    parser.add_argument('--dataset', '-ds', default='MNIST', help='dataset path.')
    parser.add_argument('--model_dir', '-md', default='./Classifiers', help='Output classifier path.')
    parser.add_argument('--filename', '-fn', default="summary.csv", help='Output csv file name.')
    parser.add_argument('--overwrite', '-ow', default=1, choices=[0, 1], type=int, help='Overwrite or not.')
    parser.add_argument('--iteration', '-it', default=0, type=int, help='Iteraction focused.')
    parser.add_argument('--impact', '-imp', type=float, default=0.05, help='Upperbound impact.')
    parser.add_argument('--threshold', '-thr', type=float, default=0.8, help='Lowerbound for validation accuracy')

    args = parser.parse_args()

    # configurations
    program_dir = args.program_dir
    dataset_dir = args.dataset
    file = args.filename  # 数据目录
    model_dir = args.model_dir  # 输出目录
    overwrite = args.overwrite
    iteraction = max(0, args.iteration)
    upper_impact = args.impact
    lower_threshold = args.threshold

    base_dir = os.path.join(program_dir, dataset_dir)
    names = [
        # "Nearest Neighbors",
        # "RBF SVM",
        # "Gaussian Process",
        # "Decision Tree",
        "Random Forest",
        # "Neural Net",
        # "AdaBoost",
        # "Naive Bayes",
        # "QDA"
    ]

    # define classifiers
    classifiers = [
        # KNeighborsClassifier(n_neighbors=len),
        # SVC(gamma=2, C=4, decision_function_shape='ovr'),
        # GaussianProcessClassifier(1.0 * RBF(1.0)),
        # DecisionTreeClassifier(criterion="gini", max_depth=5),
        RandomForestClassifier(max_depth=10, n_estimators=32, max_features=60),
        # MLPClassifier(alpha=1, max_iter=2000),
        # AdaBoostClassifier(),
        # GaussianNB(),
        # QuadraticDiscriminantAnalysis()
    ]

    # file = open("cliff_delta.txt", 'a')
    # file.write(dataset_dir + "\n")
    for cls_name, clf in zip(names, classifiers):
        start_time = time.time()

        classifier_dir = os.path.join(model_dir, dataset_dir, "normal")
        pca_dir = os.path.join(classifier_dir, "PCA")
        mi_dir = os.path.join(classifier_dir, "mutual_info")
        # wb_ori = load_workbook('{}/original.xlsx'.format(classifier_dir))
        # wb_cov = load_workbook('{}/with_coverage.xlsx'.format(classifier_dir))
        # ws_ori = wb_ori[cls_name]
        # wb_cov = wb_cov[cls_name]
        # origin = list(ws_ori.rows)
        # origin_list = []
        # for item in origin:
        #     data = [i.value for i in item]
        #     origin_list.append(data)
        # print(origin_list)
        origin = pd.read_excel('{}/original.xlsx'.format(classifier_dir), sheet_name=cls_name)
        cov = pd.read_excel('{}/with_coverage.xlsx'.format(classifier_dir), sheet_name=cls_name)
        pca = pd.read_excel('{}/with_coverage.xlsx'.format(pca_dir), sheet_name=cls_name)
        mi = pd.read_excel('{}/with_coverage.xlsx'.format(mi_dir), sheet_name=cls_name)
        # print(origin)
        # print(pca)
        # l1 = []
        # l2 = []
        # print(cls_name + "\n")
        # # file.write(cls_name + "\n")
        # for item in list(origin):
        #     print(item)
        #     vs_origin_d, size = cliffs_delta(pca[item], origin[item])
        #     vs_cov_d, _ = cliffs_delta(pca[item], cov[item])
        #     l1.append(vs_origin_d)
        #     l2.append(vs_cov_d)
        #     # print(d, size)
        #     # file.write(item + " " + str(d) + " " + size + "\n")
        # if not os.path.isfile('pca.xlsx'):
        #     wb = Workbook()
        #     ws = wb.create_sheet('pca')
        #     ws.append(['accuracy', 'macro_f1', 'micro_f1', 'auc', 'mcc'])
        # else:
        #     wb = load_workbook('pca.xlsx')
        # ws = wb['pca']
        # ws.append(l1)
        # ws.append(l2)
        # wb.save('pca.xlsx')


        l1 = []
        l2 = []
        print(cls_name + "\n")
        # file.write(cls_name + "\n")
        for item in list(origin):
            print(item)
            vs_origin_d, size = cliffs_delta(mi[item], origin[item])
            vs_cov_d, _ = cliffs_delta(mi[item], cov[item])
            l1.append(vs_origin_d)
            l2.append(vs_cov_d)
            # print(d, size)
            # file.write(item + " " + str(d) + " " + size + "\n")
        if not os.path.isfile('mutual_info.xlsx'):
            wb = Workbook()
            ws = wb.create_sheet('mi')
            ws.append(['accuracy', 'macro_f1', 'micro_f1', 'auc', 'mcc'])
        else:
            wb = load_workbook('mutual_info.xlsx')
        ws = wb['mi']
        ws.append(l1)
        ws.append(l2)
        wb.save('mutual_info.xlsx')


    # file.close()
        # print(origin)



