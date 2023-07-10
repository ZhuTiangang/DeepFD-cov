import os
import csv
from openpyxl import Workbook, load_workbook
import sys
import time
import random
import pickle
import argparse
import pandas as pd
from joblib import dump, load
from sklearn import preprocessing
from sklearn.multioutput import MultiOutputClassifier
from sklearn.metrics import f1_score
from sklearn.metrics import roc_auc_score
from sklearn.metrics import matthews_corrcoef
from sklearn import tree
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from sklearn.model_selection import train_test_split
from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF
from sklearn.tree import DecisionTreeClassifier, plot_tree, export_graphviz
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.impute import SimpleImputer


def label_encoding(raw_labels, labelEncoder=None):
    if labelEncoder is None:
        labelEncoder = preprocessing.LabelEncoder()
        labelEncoder = labelEncoder.fit(raw_labels)
    enc_label = labelEncoder.transform(raw_labels)
    # print('classes: {}'.format(', '.join(labelEncoder.classes_)))
    return labelEncoder, enc_label


def label_decoder(le, enc_label):
    return le.inverse_transform(enc_label)


def multi_label_binarizer(raw_labels, labelEncoder=None):
    if labelEncoder is None:
        labelEncoder = preprocessing.MultiLabelBinarizer()
        labelEncoder = labelEncoder.fit(raw_labels)
    enc_label = labelEncoder.transform(raw_labels)
    # print('classes: {}'.format(', '.join(labelEncoder.classes_)))
    return labelEncoder, enc_label


def train(x_train, y_train, n_features):
    clf = MultiOutputClassifier(KNeighborsClassifier(n_neighbors=n_features)).fit(x_train, y_train)
    return clf


def predict(clf, x_test):
    clf.predict(x_test)


def balance(df_all, FOCUS, n_sample=100):
    sampled = df_all[df_all[FOCUS] == "origin"].sample(n=n_sample, random_state=1)
    df_new = pd.concat([sampled, df_all[df_all[FOCUS] != "origin"]])
    print("Sample before: {}, after: {}".format(df_all.shape, df_new.shape))
    return df_new


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Multi-Lable Classifier Training.")
    parser.add_argument('--program_dir', '-pd', default='./RQ2', help='program path.')
    parser.add_argument('--dataset', '-ds', default='MNIST', help='dataset path.')
    parser.add_argument('--model_dir', '-md', default='./Classifiers', help='Output classifier path.')
    parser.add_argument('--filename', '-fn', default="summary.csv", help='Output csv file name.')
    parser.add_argument('--overwrite', '-ow', default=1, choices=[0, 1], type=int, help='Overwrite or not.')
    parser.add_argument('--iteration', '-it', default=0, type=int, help='Iteraction focused.')
    parser.add_argument('--impact', '-imp', type=float, default=0.05, help='Upperbound impact.')
    parser.add_argument('--threshold', '-thr', type=float, default=0.8, help='Lowerbound for validation accuracy')
    parser.add_argument('--with_coverage', '-cov', default=1, choices=[0, 1], type=int, help='Train with coverage or not')

    args = parser.parse_args()

    # configurations
    program_dir = args.program_dir
    dataset_dir = args.dataset
    file = args.filename        #数据目录
    model_dir = args.model_dir  #输出目录
    overwrite = args.overwrite
    iteraction = max(0, args.iteration)
    upper_impact = args.impact
    lower_threshold = args.threshold
    with_cov = args.with_coverage

    base_dir = os.path.join(program_dir, dataset_dir)

    # read csv
    df = pd.read_csv(os.path.join(base_dir, file))
    print("Read in {}, shape:{}".format(file, df.shape))

    # preprocess df
    # df = df.astype(np.float32)
    # df = df.replace([np.inf, -np.inf], np.nan)
    # df = df.fillna(0.0)

    # select samples by iteration.
    if iteraction > 0:
        df = df[df["Unnamed: 2"] == iteraction]
        print("Select Iteraction: {}. Shape: {}".format(iteraction, df.shape))
    else:
        print("Use all iteration.")

    print("Num of Fault:\n", df["num_fault"].describe())
    # df = df[df["num_fault"] >= 1]
    # print("Fault >= 1, shape:{}".format(df.shape))

    # select impact based on validate accuracy
    #df = df[df["impact_val_acc"] <= upper_impact]
    print("Upperbound for the impact on validation accuracy {}. Shape: {}".format(upper_impact, df.shape))

    # write separately because otherwise, got warning: Boolean Series key will be reindexed to match DataFrame index
    # df_slct = df[df["Unnamed: 1"] == "origin"]
    # df_slct = df_slct[df_slct["ft_val_accuracy"] < lower_threshold]
    # model_list = set(df_slct["Unnamed: 0"].unique())
    #
    # df = df[~df["Unnamed: 0"].isin(model_list)]
    # print("Lowerbound for validation accuracy {}. Shape: {}".format(lower_threshold, df.shape))

    # debug only
    # print("DataFrame:\n", df)

    # spilt features and labels
    features = list(filter(lambda x: x.startswith("ft_"), df.columns))
    if with_cov == 0:
        features = features[:148]
    # print(features)
    labels = list(filter(lambda x: x.startswith("lb_"), df.columns))

    # balance samples
    print("Balance Samples ...")

    # find the minimum group
    min_single_group = df.shape[0]
    for label in labels:
        cur_index = df[(df[label] == 1) & (df["num_fault"] == 1)].index
        min_single_group = min(len(cur_index), min_single_group)
        print("label (only this label = 1): {}, count: {}".format(label, len(cur_index)))
    min_single_group = max(50, min_single_group)
    print("Minimum: {} (may be reset).\n".format(min_single_group))

    # Balance label
    for label in labels:
        # reset index from 0
        df = df.reset_index(drop=True)

        indices = df.index
        cur_index = df[(df[label] == 1) & (df["num_fault"] == 1)].index
        print("label (only this label = 1): {}, count: {}".format(label, len(cur_index)))
        n_sample = len(cur_index) - min_single_group
        # if the number of sample has minimum
        if n_sample <= 0:
            print("no need balance for {}".format(label))
            continue

        slct_rm_index = random.sample(cur_index.tolist(), n_sample)
        slct_index = list(set(indices) - set(slct_rm_index))
        df = df.iloc[slct_index, :]
        print("after balance for {}: {}".format(label, df.shape[0]))

    # balance correct samples
    df = df.reset_index(drop=True)
    indices = df.index
    cur_index = df[df["num_fault"] == 0].index
    slct_rm_index = random.sample(cur_index.tolist(), len(cur_index) - min_single_group)
    slct_index = list(set(indices) - set(slct_rm_index))
    df = df.iloc[slct_index, :]
    print("after balance for {}: {}".format("correct", df.shape[0]))

    print("After balance:", df.shape[0])

    # ####################################
    # split features and labels
    X = df[features]
    Y = df[labels]

    # print out necessary information
    print("\nX:", X.describe())
    print("\nDescription of Labels:")
    print(Y.sum())
    print(df[["is_kill", "is_faulty"]].sum())
    print(df[df["num_fault"] >= 1]["impact_val_acc"].mean())
    print("Average Time:", df["time"].mean())
    print("AutoTrainer Identify: ", df["autoTrainer"].sum(), df[df["is_faulty"] == 1]["autoTrainer"].sum())
    
    # preprocessing X and Y
    X = X.astype(np.float32)
    Y = Y.astype(np.float32)
    X = X.replace([np.inf, -np.inf], np.nan)
    Y = Y.replace([np.inf, -np.inf], np.nan)
    X = X.fillna(0.0)
    mask_index = ~(np.max(X, axis=0) == np.min(X, axis=0))
    mask_index = mask_index.values
    X.loc[:,mask_index] = (X.loc[:,mask_index] - np.min(X.loc[:,mask_index], axis=0)) / (np.max(X.loc[:,mask_index], axis=0) - np.min(X.loc[:,mask_index], axis=0))
    X.loc[:,~mask_index] = 0.0
    # preprocessing finish

    # split train_test, test set = 0.3
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.3)
    # print(X_test[:50])
    # print(y_train[:50])


    # fill nan in X
    X_train = X_train.fillna(0.0)
    X_test = X_test.fillna(0.0)
    y_train = y_train.fillna(0.0)
    y_test = y_test.fillna(0.0)

    # encoding label
    # le, y_train = label_encoding(y_train)
    # _, y_test = label_encoding(y_test, le)
    # cls = [c for c in le.classes_]

    # necessary info

    # print info
    print("\nSplit train val set. Train: {}, Val: {}".format(X_train.shape[0], X_test.shape[0]))
    print("\nLabel Stats: {} ({})".format(len(labels), ", ".join(labels)))
    print("\nNumber of features: {}. Number of labels: {}".format(len(features), len(labels)))

    # specify classifiers' name
    names = [
        "Nearest Neighbors",
        # "RBF SVM",
        # "Gaussian Process",
        "Decision Tree",
        "Random Forest",
        "Neural Net",
        # "AdaBoost",
        # "Naive Bayes",
        # "QDA"
    ]

    # define classifiers
    classifiers = [
        KNeighborsClassifier(n_neighbors=len(labels)),
        # SVC(gamma=2, C=4, decision_function_shape='ovr'),
        # GaussianProcessClassifier(1.0 * RBF(1.0)),
        DecisionTreeClassifier(criterion="gini", max_depth=5),
        RandomForestClassifier(max_depth=10, n_estimators=32, max_features=60),
        MLPClassifier(alpha=1, max_iter=2000),
        # AdaBoostClassifier(),
        # GaussianNB(),
        # QuadraticDiscriminantAnalysis()
    ]

    print("\nStart Training Classifiers...")
    for cls_name, clf in zip(names, classifiers):
        start_time = time.time()

        classifier_dir = os.path.join(model_dir, dataset_dir)
        os.makedirs(classifier_dir, exist_ok=True)
        if with_cov == 1:
            pkl_path = os.path.join(classifier_dir, "{}_with_cov.pkl".format(cls_name))
        else:
            pkl_path = os.path.join(classifier_dir, "{}_origin.pkl".format(cls_name))
        print(pkl_path)
        if os.path.exists(pkl_path) and not overwrite:
            clf = load(pkl_path)
        else:
            try:
                clf.fit(X_train, y_train)
            except ValueError as e:
                print("- Classifier {} failed because {}".format(cls_name, e))
                continue

        # Evaluate model
        # print(dir(clf))
        # print(clf.__dict__)
        score = clf.score(X_test, y_test)
        y_pred = clf.predict(X_test)
        macro_f1 = f1_score(y_test, y_pred, average='macro')
        # print("macro f1:{}".format(macro_f1))
        micro_f1 = f1_score(y_test, y_pred, average='micro')
        # print("micro f1:{}".format(micro_f1))
        y_proba = clf.predict_proba(X_test)
        # print(np.array(y_proba).shape)
        if cls_name == 'Nearest Neighbors' or cls_name == 'Decision Tree' or cls_name == 'Random Forest':
            y_proba = np.transpose([proba[:, 1] for proba in y_proba])
            auc = roc_auc_score(y_test, y_proba, average=None)
            auc = auc.mean()
        # print(proba.shape)
        # print(np.array(y_pred).shape)
        else:
            auc = roc_auc_score(y_test, y_proba, multi_class='ovr')
        # print("auc:{}".format(auc))
        mcc = matthews_corrcoef(y_test.values.argmax(axis=1), y_pred.argmax(axis=1))
        # print("mcc:{}".format(mcc))
        print("- Classifier: {} | Score: {} | Time: {:.2f} | macro f1:{} | micro f1:{} \n auc:{} | mcc:{}".
              format(cls_name, score, time.time() - start_time, macro_f1, micro_f1, auc, mcc))
        if with_cov == 0:
            if not os.path.isfile('{}/original.xlsx'.format(classifier_dir)):
                wb = Workbook()
            else:
                wb = load_workbook('{}/original.xlsx'.format(classifier_dir))
            if cls_name not in wb.sheetnames:
                ws = wb.create_sheet(cls_name)
                ws.append(['accuracy', 'macro_f1', 'micro_f1', 'auc', 'mcc'])
            ws = wb[cls_name]
            ws.append([score, macro_f1, micro_f1, auc, mcc])
            wb.save('{}/original.xlsx'.format(classifier_dir))
        else:
            if not os.path.isfile('{}/with_coverage.xlsx'.format(classifier_dir)):
                wb = Workbook()
            else:
                wb = load_workbook('{}/with_coverage.xlsx'.format(classifier_dir))
            if cls_name not in wb.sheetnames:
                ws = wb.create_sheet(cls_name)
                ws.append(['accuracy', 'macro_f1', 'micro_f1', 'auc', 'mcc'])
            ws = wb[cls_name]
            ws.append([score, macro_f1, micro_f1, auc, mcc])
            wb.save('{}/with_coverage.xlsx'.format(classifier_dir))


        # print(clf.predict(X_test))
        # print(y_test)
        '''print(y_test.index)
        subject = []
        for i in y_test.index:
            subject.append(df["Unnamed: 0"][i])

        subject = np.array(subject)
        # print(clf.predict(X_test).shape, y_test.shape, subject.T.shape)
        result = np.column_stack((subject.T, y_test, clf.predict(X_test)))
        result = result.tolist()
        if not os.path.isfile('{}/Predict_cover.xlsx'.format(classifier_dir)):
            wb = Workbook()
        else:
            wb = load_workbook('{}/Predict_cover.xlsx'.format(classifier_dir))
        ws = wb.create_sheet(cls_name)
        for row in result:
            ws.append(row)
        wb.save('{}/Predict_cover.xlsx'.format(classifier_dir))'''





        # Save pkl
        with open(pkl_path, 'wb') as fw:
            dump(clf, fw)
        print("  - Saved to {}.".format(pkl_path))

        # draw decision tree
        # if cls_name in ["Decision Tree"]:
        #     fig, axes = plt.subplots(nrows=1, ncols=1, figsize=(10, 10), dpi=300)
        #     plot_tree(clf, feature_names=features, class_names=cls, filled=True, rounded=True, proportion=True)
        #     fig.savefig(os.path.join(base_dir, 'tree.png'))
        #     plt.show()
