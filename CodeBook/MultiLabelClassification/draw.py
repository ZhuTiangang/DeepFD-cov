import os
import csv
import tensorflow as tf
import sys
import time
import random
import pickle
import argparse
import pandas as pd
from openpyxl import Workbook, load_workbook
from joblib import dump, load
from sklearn import preprocessing
from sklearn.multioutput import MultiOutputClassifier
from cliffs_delta import cliffs_delta
from sklearn.decomposition import PCA
from sklearn.feature_selection import mutual_info_classif
from sklearn.metrics import f1_score
from sklearn.metrics import roc_auc_score
from sklearn.metrics import matthews_corrcoef
from sklearn import tree
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from sklearn.model_selection import train_test_split
from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF
from sklearn.tree import DecisionTreeClassifier, plot_tree, export_graphviz
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.impute import SimpleImputer
# from ggplot import *
from plotnine import *
import matplotlib
matplotlib.use("Agg")



def label_encoding(raw_labels, labelEncoder=None):
    if labelEncoder is None:
        labelEncoder = preprocessing.LabelEncoder()
        labelEncoder = labelEncoder.fit(raw_labels)
    enc_label = labelEncoder.transform(raw_labels)
    # print('classes: {}'.format(', '.join(labelEncoder.classes_)))
    return labelEncoder, enc_label


def label_decoder(le, enc_label):
    return le.inverse_transform(enc_label)


def multi_label_binarizer(raw_labels, labelEncoder=None):
    if labelEncoder is None:
        labelEncoder = preprocessing.MultiLabelBinarizer()
        labelEncoder = labelEncoder.fit(raw_labels)
    enc_label = labelEncoder.transform(raw_labels)
    # print('classes: {}'.format(', '.join(labelEncoder.classes_)))
    return labelEncoder, enc_label


def train(x_train, y_train, n_features):
    clf = MultiOutputClassifier(KNeighborsClassifier(n_neighbors=n_features)).fit(x_train, y_train)
    return clf


def predict(clf, x_test):
    clf.predict(x_test)


def balance(df_all, FOCUS, n_sample=100):
    sampled = df_all[df_all[FOCUS] == "origin"].sample(n=n_sample, random_state=1)
    df_new = pd.concat([sampled, df_all[df_all[FOCUS] != "origin"]])
    print("Sample before: {}, after: {}".format(df_all.shape, df_new.shape))
    return df_new


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Multi-Lable Classifier Training.")
    parser.add_argument('--program_dir', '-pd', default='./RQ2', help='program path.')
    parser.add_argument('--dataset', '-ds', default='MNIST', help='dataset path.')
    parser.add_argument('--model_dir', '-md', default='./Classifiers', help='Output classifier path.')
    parser.add_argument('--filename', '-fn', default="summary.csv", help='Output csv file name.')
    parser.add_argument('--overwrite', '-ow', default=1, choices=[0, 1], type=int, help='Overwrite or not.')
    parser.add_argument('--iteration', '-it', default=0, type=int, help='Iteraction focused.')
    parser.add_argument('--impact', '-imp', type=float, default=0.05, help='Upperbound impact.')
    parser.add_argument('--threshold', '-thr', type=float, default=0.8, help='Lowerbound for validation accuracy')

    args = parser.parse_args()

    # configurations
    program_dir = args.program_dir
    # dataset_dir = args.dataset
    file = args.filename  # 数据目录
    model_dir = args.model_dir  # 输出目录
    overwrite = args.overwrite
    iteraction = max(0, args.iteration)
    upper_impact = args.impact
    lower_threshold = args.threshold

    names = [
        "Nearest Neighbors",
        # "RBF SVM",
        # "Gaussian Process",
        "Decision Tree",
        "Random Forest",
        "Neural Net",
        # "AdaBoost",
        # "Naive Bayes",
        # "QDA"
    ]

    # define classifiers
    classifiers = [
        KNeighborsClassifier(n_neighbors=len),
        # SVC(gamma=2, C=4, decision_function_shape='ovr'),
        # GaussianProcessClassifier(1.0 * RBF(1.0)),
        DecisionTreeClassifier(criterion="gini", max_depth=5),
        RandomForestClassifier(max_depth=10, n_estimators=32, max_features=60),
        MLPClassifier(alpha=1, max_iter=2000),
        # AdaBoostClassifier(),
        # GaussianNB(),
        # QuadraticDiscriminantAnalysis()
    ]
    datasets = [
        'mnist',
        'cifar-10',
        'blob',
        'circle',
        'imdb',
        'reuters'
    ]
    for cls_name in names:
        df_gg = pd.DataFrame(columns=['dataset', 'method', 'metrics', 'value'])
        for dataset in datasets:
            classifier_dir = os.path.join(model_dir, dataset, 'normal')
            origin = pd.read_excel('{}/original.xlsx'.format(classifier_dir), sheet_name=cls_name)
            cov = pd.read_excel('{}/with_coverage.xlsx'.format(classifier_dir), sheet_name=cls_name)
            print(cls_name)
            origin['method'] = 'DeepFD'
            cov['method'] = 'DeepFD+cov'
            df = pd.concat([origin, cov])
            df['dataset'] = dataset
            # print(df)

            # print(df['accuracy'])
            for item in origin:
                if item == 'method':
                    continue
                df_cur = pd.DataFrame()
                df_cur['dataset'] = df['dataset']
                df_cur['method'] = df['method']
                df_cur[cls_name] = df[item]
                df_cur['metrics'] = item
                print(df_cur)
                df_gg = pd.concat([df_gg, df_cur])
        print(df_gg)
        df_gg['dataset'] = df_gg['dataset'].astype('category')
        df_gg['metrics'] = df_gg['metrics'].astype('category')
        df_gg['dataset'] = df_gg['dataset'].cat.reorder_categories(
            ['mnist', 'cifar-10', 'circle', 'blob', 'imdb', 'reuters'])
        df_gg['metrics'] = df_gg['metrics'].cat.reorder_categories(['accuracy', 'macro_f1', 'micro_f1', 'auc', 'mcc'])
        plot = ggplot(df_gg, aes(x='method', y=cls_name, fill='method')) + theme(
            axis_text_x=element_text(visible=False)) + geom_boxplot() + facet_grid('metrics ~ dataset', scales="free")
        # print(plot)
        plot.save('figure/' + cls_name + '.svg')
        df_gg.to_csv('figure/' + cls_name + '.csv', index=True)

    '''file = open("cliff_delta.txt", 'a')
    file.write(dataset_dir + "\n")
    for cls_name, clf in zip(names, classifiers):
        start_time = time.time()

        classifier_dir = os.path.join(model_dir, dataset_dir)
        # wb_ori = load_workbook('{}/original.xlsx'.format(classifier_dir))
        # wb_cov = load_workbook('{}/with_coverage.xlsx'.format(classifier_dir))
        # ws_ori = wb_ori[cls_name]
        # wb_cov = wb_cov[cls_name]
        # origin = list(ws_ori.rows)
        # origin_list = []
        # for item in origin:
        #     data = [i.value for i in item]
        #     origin_list.append(data)
        # print(origin_list)
        origin = pd.read_excel('{}/original.xlsx'.format(classifier_dir), sheet_name=cls_name)
        cov = pd.read_excel('{}/with_coverage.xlsx'.format(classifier_dir), sheet_name=cls_name)
        print(cls_name + "\n")
        file.write(cls_name + "\n")
        for item in list(origin):
            print(item)
            d, size = cliffs_delta(origin[item], cov[item])
            print(d, size)
            file.write(item + " " + str(d) + " " + size + "\n")

    file.close()'''
    # print(origin)
