import os
import tensorflow as tf
import numpy as np
import sys
import argparse

config = tf.compat.v1.ConfigProto(gpu_options=tf.compat.v1.GPUOptions(allow_growth=True))
sess = tf.compat.v1.Session(config=config)
from keras.models import load_model
from coverage import Coverage
import keras
import importlib
from openpyxl import Workbook, load_workbook
from keras import backend as K
sys.path.append('./data')
sys.path.append('./Utils')

def load_data(name):
    assert (name.upper() in ['MNIST', 'CIFAR-10', 'SVHN'])
    name = name.lower()
    x_train = np.load('../data/' + name + '_data/' + name + '_x_train.npy')
    y_train = np.load('../data/' + name + '_data/' + name + '_y_train.npy')
    x_test = np.load('../data/' + name + '_data/' + name + '_x_test.npy')
    y_test = np.load('../data/' + name + '_data/' + name + '_y_test.npy')
    return x_train, y_train, x_test, y_test

def get_dataset(dataset_name, base_dir=""):
    data_set = {}
    data_name = dataset_name.split('_')[0]
    data = importlib.import_module('{}'.format(data_name.lower()), package='data')

    if data_name == 'simplednn':
        choice = dataset_name.split('_')[-1]
        (x, y), (x_val, y_val) = data.load_data(method=choice)
    else:
        x = np.load('./dataset/' + data_name + '/x.npy', allow_pickle=True)
        y = np.load('./dataset/' + data_name + '/y.npy', allow_pickle=True)
        x_val = np.load('./dataset/' + data_name + '/x_val.npy', allow_pickle=True)
        y_val = np.load('./dataset/' + data_name + '/y_val.npy', allow_pickle=True)
    preprocess_func = data.preprocess
    data_set['x'] = preprocess_func(x)
    data_set['x_val'] = preprocess_func(x_val)
    if dataset_name == 'cifar10' or dataset_name == 'mnist':
        labels = 10
        data_set['y'] = keras.utils.to_categorical(y, labels)
        data_set['y_val'] = keras.utils.to_categorical(y_val, labels)
    elif data_name == 'reuters':
        labels = 46
        data_set['y'] = keras.utils.to_categorical(y, labels)
        data_set['y_val'] = keras.utils.to_categorical(y_val, labels)
    elif data_name == 'imdb':
        labels = 2
        data_set['y'] = keras.utils.to_categorical(y, labels)
        data_set['y_val'] = keras.utils.to_categorical(y_val, labels)
    else:
        data_set['y'] = y
        data_set['y_val'] = y_val
    return data_set

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Compute coverage.')
    parser.add_argument('--base', '-bs', default='./Evaluation', help='Base directory.')
    parser.add_argument('--dataset', '-ds', default='MNIST', help="Dataset.",
                        # choices=["MNIST", "MNIST2", "CIFAR-10", "Blob", "Circle", "Reuters", "IMDB"]
                        )
    parser.add_argument('--iterate', '-iter', type=int, default=1)
    args = parser.parse_args()
    parent_dir = args.base
    dataset_name = args.dataset
    dir = (parent_dir + '/' + dataset_name + '/normal')
    i = args.iterate


    # x_train, y_train, x_test, y_test = load_data(dataset)
    is_dataset_loaded = False
    # print(os.listdir())

    if is_dataset_loaded:
        print("Dataset already {} loaded.".format(dataset_name))
    else:
        dataset = get_dataset(dataset_name)
        is_dataset_loaded = True
        print("Dataset {} loaded.".format(dataset_name))

    # dir = ('./Evaluation/' + 'blob' + '/normal')

    for model in os.listdir(dir):
        if os.path.isdir(os.path.join(dir, model)):
            print("\nModel : {}".format(model))
            model_path = os.path.join(dir, model)
        else:
            continue
        for fault in os.listdir(model_path):


            print('\nfault : {}'.format(fault))
            fault_path = os.path.join(model_path, fault)
            if os.path.isdir(fault_path):
                # for i in iter:
                result_dir = (fault_path + '/iter_{}'.format(i) + '/result_dir')
                print(result_dir)
                if os.path.isdir(result_dir) and not os.path.isfile('{}/Coverage_result.xlsx'.format(result_dir)):
                    # os.chdir(result_dir)
                    epoches = [os.path.join(result_dir, f) for f in os.listdir(result_dir) if f.endswith("hdf5")]
                    trend_train_nc1 = []
                    trend_train_nc2 = []
                    trend_train_nc3 = []
                    trend_train_nc4 = []
                    trend_train_nc5 = []
                    trend_train_tknc = []
                    trend_train_kmnc = []
                    trend_train_tknp = []
                    trend_train_nbc = []
                    trend_train_snac = []
                    for epoch in epoches:
                        print(epoch)
                        model = load_model(epoch)
                        model.summary()
                        model_layer = len(model.layers) - 1
                        l = []
                        # model.summary()
                        for j in range(1, model_layer):
                            if model.layers[j].output.shape[1] != None:
                                l.append(j)
                                #print(l)
                        coverage = Coverage(model, dataset['x'], dataset['x'])
                        nc1, nc2, nc3, nc4, nc5, kmnc, nbc, snac, tknc, tknp = coverage.all(l)
                        trend_train_nc1.append(nc1)
                        trend_train_nc2.append(nc2)
                        trend_train_nc3.append(nc3)
                        trend_train_nc4.append(nc4)
                        trend_train_nc5.append(nc5)
                        trend_train_tknc.append(tknc)
                        trend_train_kmnc.append(kmnc)
                        trend_train_tknp.append(tknp)
                        trend_train_nbc.append(nbc)
                        trend_train_snac.append(snac)
                        K.clear_session()

                    train_trend = [['NC0.1'] + trend_train_nc1,
                                   ['NC0.3'] + trend_train_nc2,
                                   ['NC0.5'] + trend_train_nc3,
                                   ['NC0.7'] + trend_train_nc4,
                                   ['NC0.9'] + trend_train_nc5,
                                   ['TKNC'] + trend_train_tknc,
                                   ['TKNP'] + trend_train_tknp,
                                   ['KMNC'] + trend_train_kmnc,
                                   ['NBC'] + trend_train_nbc,
                                   ['SNAC'] + trend_train_snac]

                    if not os.path.isfile('{}/Coverage_result.xlsx'.format(result_dir)):
                        wb = Workbook()
                    else:
                        wb = load_workbook('{}/Coverage_result.xlsx'.format(result_dir))
                    ws = wb.create_sheet('iter_{0}'.format(i))
                    for row in train_trend:
                        ws.append(row)
                    wb.save('{}/Coverage_result.xlsx'.format(result_dir))
